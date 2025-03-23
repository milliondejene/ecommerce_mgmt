from django.db import models
from django.db.models import Sum, F
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

# Create your models here.

class Product(models.Model):
    name = models.CharField(max_length=200)
    sku = models.CharField(max_length=50, unique=True)
    unit_price = models.DecimalField(max_digits=10, decimal_places=2)

    def __str__(self):
        return f"{self.name} ({self.sku})"

class PurchaseOrder(models.Model):
    STATUS_CHOICES = [
        ('PENDING', 'Pending'),
        ('COMPLETED', 'Completed'),
        ('CANCELED', 'Canceled'),
    ]
    
    vendor = models.CharField(max_length=200)
    order_date = models.DateField()
    status = models.CharField(max_length=20, choices=STATUS_CHOICES)
    products = models.ManyToManyField(Product, through='PurchaseOrderLineItem')

    def __str__(self):
        return f"PO-{self.id} - {self.vendor}"

class PurchaseOrderLineItem(models.Model):
    purchase_order = models.ForeignKey(PurchaseOrder, on_delete=models.CASCADE)
    product = models.ForeignKey(Product, on_delete=models.CASCADE)
    quantity = models.IntegerField()
    cost = models.DecimalField(max_digits=10, decimal_places=2)

class Invoice(models.Model):
    STATUS_CHOICES = [
        ('UNPAID', 'Unpaid'),
        ('PAID', 'Paid'),
        ('OVERDUE', 'Overdue'),
    ]
    
    customer = models.CharField(max_length=200)
    invoice_date = models.DateField()
    status = models.CharField(max_length=20, choices=STATUS_CHOICES)
    products = models.ManyToManyField(Product, through='InvoiceLineItem')

    def __str__(self):
        return f"INV-{self.id} - {self.customer}"

    def get_total(self):
        return self.invoicelineitem_set.annotate(
            line_total=F('quantity') * F('price_each')
        ).aggregate(
            total=Sum('line_total')
        )['total'] or 0
        
    def export_to_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        centered = Alignment(horizontal="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Invoice Details
        ws['A1'] = "Invoice Details"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        details = [
            ("Invoice Number:", f"INV-{self.id}"),
            ("Customer:", self.customer),
            ("Date:", self.invoice_date.strftime("%Y-%m-%d")),
            ("Status:", self.status),
        ]
        
        for i, (label, value) in enumerate(details, start=2):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Products Table
        row = 7
        headers = ["Product", "SKU", "Quantity", "Price Each", "Total"]
        for col, header in enumerate(headers, start=1):
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 15
            
        row += 1
        items = self.invoicelineitem_set.select_related('product').all()
        
        for item in items:
            ws[f'A{row}'] = item.product.name
            ws[f'B{row}'] = item.product.sku
            ws[f'C{row}'] = item.quantity
            ws[f'D{row}'] = float(item.price_each)
            ws[f'E{row}'] = float(item.quantity * item.price_each)
            
            for col in range(1, 6):
                cell = ws[f'{get_column_letter(col)}{row}']
                cell.border = border
                cell.alignment = centered
            row += 1
            
        # Total
        ws[f'D{row+1}'] = "Total:"
        ws[f'D{row+1}'].font = Font(bold=True)
        ws[f'E{row+1}'] = float(self.get_total())
        ws[f'E{row+1}'].font = Font(bold=True)
        
        # Ensure exports directory exists
        os.makedirs('exports', exist_ok=True)
        
        # Save file
        filename = f"exports/INV-{self.id}_{self.customer}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename

class InvoiceLineItem(models.Model):
    invoice = models.ForeignKey(Invoice, on_delete=models.CASCADE)
    product = models.ForeignKey(Product, on_delete=models.CASCADE)
    quantity = models.IntegerField()
    price_each = models.DecimalField(max_digits=10, decimal_places=2)
